"""--------------------------------------------------
createUpdateQuery.py

    csvからUPDATEクエリ付きのEXCELファイルを作成する。
    @in:
        *.csv
        settings.update.json
    @out: 任意名.xlsx
--------------------------------------------------"""
# 外部ライブラリ
import shutil
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from pathlib import Path
# カスタム関数
import functions as cf
# クラス
from modules import UploadManager


'''========================================
設定ファイル読み込み
========================================'''
jsn = 'settings.update.json'

try:
    inputNum = int(input(f'{jsn}のキー番号を選んでください: '))
except ValueError:
    # Enter an integer: abc
    exit(f'input error: {jsn}のキー番号を確認してください。')


'''========================================
前準備
========================================'''
params = cf.getParamsByJson(str(inputNum), jsn)

root = Path('.')
input_dir = root / params['input_dir']
input_files = list(input_dir.glob('**/*.csv'))
output_dir = root / params['output_dir']
output_file = f'{output_dir}/{params['output_file']}'

table = params['table']
key_val_dict = params['update_key_val']

# 特定のテーブル用の処理
if inputNum == 1:
    um = UploadManager(input_files)
    key_val_dict['upload_file_url'] = \
        um.getUrlByFiles(params['replace_domain'])
    key_val_dict['upload_filename'] = \
        um.getFileNameByUrls(key_val_dict['upload_file_url'])

# 入力元: 存在チェック
try:
    if not input_dir.exists():
        raise FileNotFoundError
except FileNotFoundError:
    exit(f'{input_dir} が存在しません。')

try:
    if not input_files:
        raise FileNotFoundError
except FileNotFoundError:
    exit(f'{input_dir} にcsvファイルが存在しません。')

# 出力先: 初期化
if output_dir.exists():
    shutil.rmtree(output_dir)
output_dir.mkdir()


'''========================================
csv >> excel書き出し
========================================'''

ws_list = []

for f in input_files:
    df_head = pd.read_csv(f, header=None, nrows=1)

for f_idx, f in enumerate(input_files):

    offset_num = cf.applyOffsetNum(table)
    df = pd.read_csv(f, header=offset_num)
    df_trimmed = df.replace(
        r'(^[\'|\"|\s]{1}[\s|\t]*|[\s|\t]*[\'|\"]{1}$)', '', regex=True)

    df_result = cf.duplicateDf(df_trimmed, params['sortkey'])

    # ファイル名をシート名にする
    ws_title = re.sub('.csv', '', cf.getFileName(f))
    ws_list.append(ws_title)

    if f_idx == 0:
        df_result.to_excel(output_file, sheet_name=ws_title, index=False)
    else:
        with pd.ExcelWriter(
            output_file,
            engine='openpyxl',
            mode='a',
            if_sheet_exists='replace'
        ) as writer:
            df_result.to_excel(writer, sheet_name=ws_title, index=False)


'''========================================
excel処理
========================================'''

wb = load_workbook(output_file)

key_idxs = cf.getColumnIndex(
    list(df_result.columns), key_val_dict.keys())
keys = cf.getColumnNames(key_val_dict)
vals_per_ws = cf.getUpdtSrc(ws_list, key_val_dict)

for ws_idx, ws in enumerate(ws_list):
    ws = wb[ws]

    # 更新用のカラム、値情報を配置するためのセル番地を格納
    key_addrs = []
    val_addrs = []
    for i in range(1, len(keys) + 1):
        key_addrs.append(ws.cell(row=i, column=1).coordinate)
        val_addrs.append(ws.cell(row=i, column=2).coordinate)

    fill_color1 = PatternFill(fgColor='7AF5D8', fill_type='solid')
    fill_color2 = PatternFill(fgColor='F5E7EE', fill_type='solid')

    emphasis_font_color = Font(color='FF0000')

    # 更新カラム分だけ上から行追加していく。A列にカラム名、B列に値を設定する。
    for k_idx, key in enumerate(keys):
        ws.insert_rows(k_idx + 1)
        ws[key_addrs[k_idx]].value = key
        ws[key_addrs[k_idx]].fill = fill_color1
        ws[val_addrs[k_idx]].value = vals_per_ws[ws_idx][k_idx]

    # 追加した行数分下から表ループ開始
    start_row = len(key_addrs) + 2
    interval = 2

    # interval間隔で表を走査
    for i in range(start_row, ws.max_row+1, interval):

        # 新規行にあたる行のセルを走査する
        for row in ws.iter_rows(min_row=i, max_row=i):
            for cell in row:
                cell.fill = fill_color2

                # 現在のセルが更新キー列か判定 >> true: 数式埋め込み( 更新値情報を持つセルを参照させる )
                # key_idxs: 更新対象となるキー列のインデックス
                # val_addrs: 参照元とする更新値情報を持つセル番地
                # キー列番号とセル番地(list)の長さ・序列は対応している
                for k_idx, key in enumerate(key_idxs):
                    if cell.column == key:
                        cell.value = f'={val_addrs[k_idx]}'
                        cell.font = emphasis_font_color

    '''------------------------------
    SQL生成スタート
    ------------------------------'''
    # 表の最後にSQL用の列を追加
    append_col_no = ws.max_column+1

    # SET句で使うカラム(セル番地)を用意
    set_key_addrs = []
    keys_info = list(key_val_dict.keys())
    for row in ws.iter_rows(min_row=start_row-1, max_row=start_row-1):
        for cell in row:
            for key in keys_info:
                if cell.value == key:
                    set_key_addrs.append(cell.coordinate)

    '''********************
    UPDATE句
    ********************'''
    row_count = 1
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=ws.max_row,
        min_col=append_col_no,
        max_col=append_col_no
    ):
        # sql_head 切り戻し用の行はコメントアウトしておく()
        if row_count % 2 == 0:
            sql_u_head = f'="-- UPDATE `{table}` SET"'
        else:
            sql_u_head = f'="UPDATE `{table}` SET"'
        row_count += 1

        sql_u_body = ''
        sql_u_cond = '&"WHERE"&'

        # UPDATE SET に続く[ カラム=値 ]をループで生成
        for cell in row:
            # query_body
            for k_idx, k_num in enumerate(key_idxs):
                set_key_addr = set_key_addrs[k_idx]
                set_val_addr = ws.cell(row=cell.row, column=k_num).coordinate
                value = ws[set_val_addr].value

                # ループ 1st: ループ回数判定(最終回か否か)
                # ループ 2nd: NULL判定
                if k_idx == len(key_idxs) - 1:
                    if re.fullmatch('NULL', value, flags=re.IGNORECASE):
                        sql_u_body += \
                            f'&" `"&{set_key_addr}&"` = "&{set_val_addr}&" "'
                    else:
                        sql_u_body += \
                            f'&" `"&{set_key_addr}&"` = \'"&{
                                set_val_addr}&"\' "'
                else:
                    if re.fullmatch('NULL', value, flags=re.IGNORECASE):
                        sql_u_body += \
                            f'&" `"&{set_key_addr}&"` = "&{set_val_addr}&","'
                    else:
                        sql_u_body += \
                            f'&" `"&{set_key_addr}&"` = \'"&{
                                set_val_addr}&"\',"'

            # query_condition
            cond_key_addr = ws.cell(row=start_row-1, column=1).coordinate
            cond_val_addr = ws.cell(row=cell.row, column=1).coordinate

            sql_u_cond += \
                f'" `"&{cond_key_addr}&"` = \'"&{cond_val_addr}&"\';"'

            # cell.valueに埋め込み
            cell.value = (sql_u_head + sql_u_body + sql_u_cond)

    '''********************
    SELECT句
    ********************'''
    s_cond_key = ws.cell(row=start_row-1, column=1).coordinate
    s_cond_vals = []

    # 1列目を走査し、条件カラム値を取得
    for i in range(start_row, ws.max_row):

        for row in ws.iter_rows(min_row=i, max_row=i, min_col=1, max_col=1):
            for cell in row:
                s_cond_vals.append(cell.value)

    s_condition_vals = sorted(list(set(s_cond_vals)))

    sql_s_head = f'= "SELECT * FROM `{table}` WHERE `"&{
        s_cond_key}&"` IN('

    sql_s_body = ''
    for idx, val in enumerate(s_condition_vals):
        if idx == len(s_condition_vals)-1:
            sql_s_body += f' \'{val}\' );"'
        else:
            sql_s_body += f' \'{val}\', '

    ws.cell(row=ws.max_row+2, column=1).value = '確認用クエリ'
    ws.cell(row=ws.max_row+1, column=1).value = sql_s_head + sql_s_body

    # ヘッダー行にフィルター設定
    ws.auto_filter.ref = f'{ws.cell(row=start_row - 1, column=1).coordinate}:{
        ws.cell(row=start_row - 1, column=ws.max_column - 1).coordinate}'

wb.save(output_file)
